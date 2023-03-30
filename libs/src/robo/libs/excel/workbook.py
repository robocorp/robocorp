import logging
import pathlib
from typing import Any, Optional, Union
from typing_extensions import Literal

from openpyxl.utils.exceptions import InvalidFileException
from robo.libs.excel._worksheet import Worksheet
from robo.libs.excel._workbooks import XlsWorkbook, XlsxWorkbook


def _load_workbook(
    path: str, data_only: bool, read_only: bool
) -> Union[XlsWorkbook, XlsxWorkbook]:
    # pylint: disable=broad-except
    parsed_path = pathlib.Path(path).resolve(strict=True)

    try:
        book = XlsxWorkbook(parsed_path)
        book.open(data_only=data_only, read_only=read_only)
        return book
    except InvalidFileException as exc:
        logging.debug(exc)  # Unsupported extension, silently try xlrd
    except Exception as exc:
        logging.info("Failed to open as Office Open XML (.xlsx) format: %s", exc)

    try:
        book = XlsWorkbook(parsed_path)
        book.open()
        return book
    except Exception as exc:
        logging.info("Failed to open as Excel Binary Format (.xls): %s", exc)

    raise ValueError(
        f"Failed to open Excel file ({path}), "
        "verify that the path and extension are correct"
    )


def create_workbook(filename: str, fmt: Literal["xlsx", "xls"] = "xlsx"):
    # Creates a new workbook
    # files.create_workbook()
    return Workbook()


def open_workbook(filename: str):
    # Opens an existing workbook
    # files.open_workbook()
    return Workbook()


class Workbook:
    def __init__(self):
        # Internal API, for users there is create_ and open_ workbook functions
        self._excel = None

    def save(self):
        # files.save_workbook()
        pass

    def close(self):
        # Could also be a context manager and auto close
        # files.close_workbook()
        pass

    def worksheet(self, name=None):
        """If name is not provided take the first worksheet?"""
        return Worksheet()

    def create_worksheet(
        self,
        name: str,
        content: Optional[Any] = None,
        exist_ok: Optional[bool] = False,
        header: Optional[bool] = False,
    ):
        return Worksheet()


# these may be removed, replaced by handling worksheets yourself? or these will go to workbook!
# files.set_active_worksheet()
# files.create_worksheet()
# files.remove_worksheet()
# files.worksheet_exists()
# files.list_worksheets()
# files.get_active_worksheet()
