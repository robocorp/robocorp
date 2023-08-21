import logging
from pathlib import Path
from typing import Literal, Optional, Union

from openpyxl.utils.exceptions import InvalidFileException
from robocorp.excel._tables import Table
from robocorp.excel._types import PathType
from robocorp.excel._workbooks import XlsWorkbook, XlsxWorkbook
from robocorp.excel._workbook import Workbook

__version__ = "0.4.0"
version_info = [int(x) for x in __version__.split(".")]

LOGGER = logging.getLogger(__name__)


def create_workbook(
    fmt: Literal["xlsx", "xls"] = "xlsx",
    sheet_name: Optional[str] = None,
) -> Workbook:
    """Create and open a new Excel workbook in memory.

    Automatically also creates a new worksheet with the name ``sheet_name``.

    **Note:** Use the ``save`` method to store the workbook into file.

    Args:
        fmt: The file format for the workbook. Supported file formats: ``xlsx``, ``xls``.
        sheet_name: The name for the initial sheet. If None, then set to ``Sheet``.

    Returns:
        Workbook: The created Excel workbook object.

    Example:
        .. code-block:: python

            workbook = create_workbook("xlsx", sheet_name="Sheet1")

    """
    # FIXME: add missing types from this docs page https://support.microsoft.com/en-us/office/file-formats-that-are-supported-in-excel-0943ff2c-6014-4e8d-aaea-b83d51d46247
    # check which of these our python lib behind scenes supports: xlsx, xls, xlsm, xltm, xltx, xlt, xlam, xlsb, xla, xlr, xlw, xll
    # removed path, as it is only used when saved
    # files.create_workbook()
    if fmt == "xlsx":
        workbook = XlsxWorkbook()
    elif fmt == "xls":
        workbook = XlsWorkbook()
    else:
        raise ValueError(f"Unknown format: {fmt}")

    workbook.create()
    if sheet_name is not None:
        workbook.rename_worksheet(sheet_name, workbook.active)

    return Workbook(workbook)


def open_workbook(
    path: PathType,
    data_only: bool = False,
    read_only: bool = False,
) -> Workbook:
    """Open an existing Excel workbook.

    Opens the workbook in memory.
    The file can be in either ``.xlsx`` or ``.xls`` format.

    Args:
        path: path to Excel file
        data_only: controls whether cells with formulas have either
         the formula (default, False) or the value stored the last time Excel
         read the sheet (True). Affects only ``.xlsx`` files.

    Returns:
        Workbook: Workbook object

    Example:

        ::

            # Open workbook with only path provided
            workbook = open_workbook("path/to/file.xlsx")

            # Open workbook with path provided and reading formulas in cells
            # as the value stored
            # Note: Can only be used with XLSX workbooks
            workbook = open_workbook("path/to/file.xlsx", data_only=True)

    """
    parsed_path = Path(path).resolve(strict=True)

    try:
        book = XlsxWorkbook()
        book.open(parsed_path, data_only=data_only, read_only=read_only)
        return book
    except InvalidFileException as exc:
        logging.debug(exc)  # Unsupported extension, silently try xlrd
    except Exception as exc:
        logging.info("Failed to open as Office Open XML (.xlsx) format: %s", exc)

    if data_only or read_only:
        logging.warning(
            ".xls workbooks don't support 'data_only' or 'read_only' options"
        )

    try:
        book = XlsWorkbook(parsed_path)
        book.open(parsed_path)
        return book
    except Exception as exc:
        logging.info("Failed to open as Excel Binary Format (.xls): %s", exc)

    raise ValueError(
        f'Failed to open Excel file ("{path}"), verify that the path and extension are correct'
    )

    return Workbook(_load_workbook(path, data_only, read_only))


__all__ = [
    "create_workbook",
    "open_workbook",
    "Table",
]
